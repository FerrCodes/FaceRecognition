import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import os
import io

def convert_to_excel_with_photos():
    # Baca hasil deteksi
    with open("hasil_deteksi.json", "r", encoding="utf-8") as f:
        data = json.load(f)
    
    # Konversi ke DataFrame
    df = pd.DataFrame(data)
    
    # Simpan ke Excel dulu (tanpa foto)
    excel_file = "hasil_deteksi_dengan_foto.xlsx"
    df.to_excel(excel_file, index=False)
    
    # Load workbook untuk menambahkan foto
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Sesuaikan lebar kolom
    ws.column_dimensions['A'].width = 40  # File
    ws.column_dimensions['B'].width = 20  # Status
    ws.column_dimensions['C'].width = 20  # Jarak
    ws.column_dimensions['D'].width = 25  # Waktu
    ws.column_dimensions['E'].width = 15  # Foto (kolom baru)
    
    # Tambahkan header untuk kolom foto
    ws['E1'] = 'Foto'
    
    # Proses setiap baris untuk menambahkan foto
    FOLDER_FOTO = "foto_dari_hp"
    
    for idx, row in enumerate(data, start=2):  # start=2 karena baris 1 adalah header
        filename = row.get('file', '').strip()
        foto_path = os.path.join(FOLDER_FOTO, filename)
        
        if os.path.exists(foto_path):
            try:
                # Resize foto agar tidak terlalu besar di Excel
                img = PILImage.open(foto_path)
                img.thumbnail((100, 100))  # Maks 100x100 pixel
                
                # Simpan ke buffer
                img_buffer = io.BytesIO()
                img.save(img_buffer, format='JPEG')
                img_buffer.seek(0)
                
                # Buat image object untuk openpyxl
                excel_img = Image(img_buffer)
                
                # Sesuaikan tinggi baris
                ws.row_dimensions[idx].height = 80
                
                # Anchor image ke cell E
                cell = f'E{idx}'
                excel_img.anchor = cell
                
                # Tambahkan image ke worksheet
                ws.add_image(excel_img)
                
            except Exception as e:
                print(f"Gagal menambahkan foto {filename}: {e}")
                ws[f'E{idx}'] = 'Foto error'
        else:
            ws[f'E{idx}'] = 'File tidak ada'
    
    # Simpan workbook
    wb.save(excel_file)
    print(f"✅ File {excel_file} berhasil dibuat dengan foto!")
    print("📁 Buka file tersebut dengan Microsoft Excel atau LibreOffice Calc")

if __name__ == "__main__":
    convert_to_excel_with_photos()